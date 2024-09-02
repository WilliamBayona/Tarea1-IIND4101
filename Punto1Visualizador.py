from openpyxl import*
from gurobipy import*
import aspose.pdf as ap

# Carga de Parametros
book = load_workbook("Punto1Datos.xlsx")

beneficiosSheet=book["Beneficios"]
comunicacionesSheet=book["Comunicaciones"]
costosSheet=book["Costos"]


#Definicion de Parametros
b = {} #Beneficios
for fila in range(2,6):
    for columna in range (2,8):
        sede = beneficiosSheet.cell(fila,1).value
        facultad = beneficiosSheet.cell(1,columna).value
        b[(facultad,sede)]=beneficiosSheet.cell(fila,columna).value

b= {key: value * 1000000 for key,value in b.items()}
#print(b)
        
c = {} #Comunicaciones
for fila in range(2,8):
    for columna in range (2,8):
        facultad = comunicacionesSheet.cell(fila,1).value
        facultad2 = comunicacionesSheet.cell(1,columna).value
        if comunicacionesSheet.cell(fila,columna).value != "-":
          c[(facultad,facultad2)]=comunicacionesSheet.cell(fila,columna).value  
          
c= {key: value * 1000 for key,value in c.items()}
#print(c)

k = {} #Costos Comunicacion
for fila in range(2,6):
    for columna in range (2,6):
        sede = costosSheet.cell(fila,1).value
        sede2 =costosSheet.cell(1,columna).value
        if costosSheet.cell(fila,columna).value != "-":
          k[(sede,sede2)]=costosSheet.cell(fila,columna).value 
          
k= {key: value * 1000 for key,value in k.items()}
#print(k)



#Conjuntos
S = list() #Sedes
for fila in range(2,6):
    sede = beneficiosSheet.cell(fila,1).value
    S.append(sede)
#print(S)

F = list() #Facultades
for columna in range(2,8):
    facultad = beneficiosSheet.cell(1,columna).value
    F.append(facultad)
#print(F)



#Modelo de Optimización
m = Model("Punto1")

#Definicion de variables de decisión
x = m.addVars(F,S,vtype=GRB.BINARY, name = "x")

y = m.addVars(F,S,F,S,vtype=GRB.CONTINUOUS, name = "y")

#Restricciones del problema

#1. Todas las facutades se encuentran en una sede
for f in F:
    m.addConstr(quicksum(x[f,s] for s in S) == 1)

#2. Todas las sedes tienen al menos una facultad
for s in S:
    m.addConstr(quicksum(x[f,s] for f in F) >= 1)
    
#3. Todas las sedes tienen maximo tres facultades
for s in S:
    m.addConstr(quicksum(x[f,s] for f in F) <= 3)
    
#4.	La variable auxiliar yfsae siempre será menor o igual que xfs
for s in S:
    for f in F:
        for e in S:
            for a in F:
                m.addConstr(y[f,s,a,e] <= x[f,s])


#5.	La variable auxiliar yfsae siempre será menor o igual que xae
for s in S:
    for f in F:
        for e in S:
            for a in F:
                m.addConstr(y[f,s,a,e] <= x[a,e])
                

#6.	Restricción que cubre el ultimo caso de la tabla de verdad
for s in S:
    for f in F:
        for e in S:
            for a in F:
                m.addConstr(y[f,s,a,e] >= x[f,s]+x[a,e]-1)
        
    

#Funcion Objetivo
m.setObjective(quicksum(x[f,s]*b[f,s] for f in F for s in S)
               -quicksum(y[f,s,a,e]*c[f,a]*k[s,e] for f in F for s in S for a in F for e in S),GRB.MAXIMIZE)

#print(quicksum(x[f,s]*b[f,s] for f in F for s in S))
#
m.update()
m.setParam("Outputflag",0)
m.optimize()
z = m.getObjective().getValue()

#Imprimir resultados en consola
z  = "${:,.2f}".format(z)
print(z)

for f,s, in x.keys():
    if x[f,s].x>0:
        print("Facultad ",f," en la sede ",s)
    

    
# Cargar el documento PDF preexistente
pdf = ap.Document("plantilla.pdf")

# Diccionario para almacenar las facultades asignadas a cada sede
sedes_facultades = {
    'Norte': [],
    'Oriente': [],
    'Occidente': [],
    'Sur': []
}

# Llenar el diccionario con las facultades correspondientes a cada sede
for f, s in x.keys():
    if x[f, s].x > 0:
        sedes_facultades[s].append(f)

# Función para reemplazar el texto en el PDF
def reemplazar_texto(original, nuevo_texto):
    txt_absorber = ap.text.TextFragmentAbsorber(original)
    pdf.pages.accept(txt_absorber)
    
    # Reemplazar el texto encontrado
    for txt_fragment in txt_absorber.text_fragments:
        txt_fragment.text = nuevo_texto

# Reemplazar cada casilla con la facultad correspondiente
for sede, facultades in sedes_facultades.items():
    for i in range(1, 4):  # Solo hay 3 campos por sede
        original_texto = f"{sede.lower()}{i}"
        nuevo_texto = f"Facultad {facultades[i - 1]}" if i <= len(facultades) else ""
        reemplazar_texto(original_texto, nuevo_texto)

# Reemplazar el valor de la FO en el campo "valor"
reemplazar_texto("valor", str(z))

# Guardar el PDF actualizado
pdf.save("informe.pdf")
    
    
