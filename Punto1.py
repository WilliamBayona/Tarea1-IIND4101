from openpyxl import*
from gurobipy import*

# Carga de Parametros
book = load_workbook("Punto1Datos.xlsx")

beneficiosSheet=book["Beneficios"]
comunicacionesSheet=book["Comunicaciones"]
costosSheet=book["Costos"]


#Definicion de Parametros
b = {} #Beneficios
for fila in range(2,6):
    for columna in range (2,8):
        sede = beneficiosSheet.cell(fila,1).value
        facultad = beneficiosSheet.cell(1,columna).value
        b[(facultad,sede)]=beneficiosSheet.cell(fila,columna).value
#print(b)
        
c = {} #Comunicaciones
for fila in range(2,8):
    for columna in range (2,8):
        facultad = comunicacionesSheet.cell(fila,1).value
        facultad2 = comunicacionesSheet.cell(1,columna).value
        if comunicacionesSheet.cell(fila,columna).value != "-":
          c[(facultad,facultad2)]=comunicacionesSheet.cell(fila,columna).value  
#print(c)

k = {} #Costos Comunicacion
for fila in range(2,6):
    for columna in range (2,6):
        sede = costosSheet.cell(fila,1).value
        sede2 =costosSheet.cell(1,columna).value
        if costosSheet.cell(fila,columna).value != "-":
          k[(sede,sede2)]=costosSheet.cell(fila,columna).value 
#print(k)



#Conjuntos
S = list() #Sedes
for fila in range(2,6):
    sede = beneficiosSheet.cell(fila,1).value
    S.append(sede)
#print(S)

F = list() #Facultades
for columna in range(2,8):
    facultad = beneficiosSheet.cell(1,columna).value
    F.append(facultad)
#print(F)



#Modelo de Optimización
m = Model("Punto1")

#Definicion de variables de decisión
x = m.addVars(F,S,vtype=GRB.BINARY, name = "x")

#Restricciones del problema

#Todas las facutades se encuentran en una sede
for f in F:
    m.addConstr(quicksum(x[f,s] for s in S) == 1)

#Todas las sedes tienen al menos una facultad
for s in S:
    m.addConstr(quicksum(x[f,s] for f in F) >= 1)
    
#Todas las sedes tienen maximo tres facultades
for s in S:
    m.addConstr(quicksum(x[f,s] for f in F) <= 3)
    
    

#Funcion Objetivo
m.setObjective(quicksum(x[f,s]*b[f,s] for f in F for s in S)-quicksum(x[f,s]*x[a,e]*c[f,a]*k[s,e] for f in F for s in S for a in F for e in S),GRB.MAXIMIZE)

#
m.update()
m.setParam("Outputflag",0)
m.optimize()
z = m.getObjective().getValue()

#Imrpimir resultados en consola
print(z)

for f,s, in x.keys():
    if x[f,s].x>0:
        print("Facultad ",f," en la sede ",s)
    
    
    
